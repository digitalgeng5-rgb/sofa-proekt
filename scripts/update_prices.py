# -*- coding: utf-8 -*-
"""
Обновление цен в партнёрских фидах по новому прайсу GENGLASS 2026.

Правила (согласованы с заказчиком):
  * менять ТОЛЬКО цены, подсвеченные в прайсе зелёным (заливка #D9EAD3);
  * в самих фидах не менять ничего, кроме значения цены;
  * матчинг по артикулу с нормализацией: trim, схлопывание пробелов (вкл. NBSP),
    UPPER, срез ведущих нулей;
  * округление не применяется — цена переносится как в прайсе;
  * позиции, которых нет в прайсе, остаются со старой ценой и выносятся отдельным листом.

Особенность прайса: одна строка = одна модель, а в колонке «Артикул» перечислено
несколько артикулов через перевод строки — все они делят одну цену «ЕРЦ».

Запуск:  python3 scripts/update_prices.py
"""
import os, re, shutil, warnings
warnings.filterwarnings('ignore')

import pandas as pd
import xlrd, openpyxl
from xlutils.copy import copy as xlcopy
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FEEDS_DIR = 'Feeds claude'
OUT_DIR = 'out'
PRICE_FILE = 'Прайс GENGLASS 2026 дилерский.xlsx'
GREEN = 'FFD9EAD3'          # заливка «обновлённая цена» в прайсе
PRICE_HEADER = 'ерц'        # колонка цены в прайсе
ART_HEADER = 'артикул'

# Реестр фидов (структура выверена по файлам). Индексы 0-based.
FEEDS = [
    dict(partner='Genglass / Basicdecor',
         file='Genglass_фиды_Инфографика_Basicdecor_5.08.xls',
         sheet='Выгрузка товаров', kind='xls', sheet_idx=0,
         header=1, article=4, price=6,
         article_name='Артикул поставщика', price_name='Розничная цена'),
    dict(partner='InMyRoom',
         file='InMyRoom 13.07.xls',
         sheet='Лист1', kind='xls', sheet_idx=0,
         header=0, article=14, price=3,
         article_name='vendorCode', price_name='price'),
    dict(partner='MEBHOME',
         file='Рабочий файл 2025 MEBHOME (10) (1) (3) (1).xlsx',
         sheet='Лист3', kind='xlsx',
         header=0, article=4, price=5,
         article_name='Артикул', price_name='РРЦ с доставкой до 15км'),
]
SKIPPED = [dict(partner='М&M', file='М&M (1) (6).xlsx',
                reason='В файле нет колонки с артикулом — матчинг по артикулу неприменим. '
                       'Файл оставлен без изменений (согласовано).')]

# Кириллические буквы, визуально неотличимые от латинских. Используются ТОЛЬКО
# для диагностики (лист «Кириллица в артикуле»), в матчинге не применяются.
CYR_TO_LAT = str.maketrans('АВЕКМНОРСТУХІЁЗ', 'ABEKMHOPCTYXIEE')

ST_UPDATED = 'Обновлена по зелёной цене'
ST_GREEN_SAME = 'Зелёная, но цена уже совпадала'
ST_NOT_GREEN = 'Есть в прайсе, но не зелёная — не трогали'
ST_ABSENT = 'Нет в прайсе — оставлена старая цена'
ST_NOPRICE = 'Нет старой цены — не трогали'


def norm_article(v):
    """trim -> схлопывание пробелов (вкл. NBSP/узкий пробел) -> UPPER -> срез ведущих нулей."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).replace('\xa0', ' ').replace(' ', ' ').replace(' ', ' ').strip()
    if s.lower() in ('', 'nan', 'none'):
        return None
    s = re.sub(r'\s+', ' ', s).upper().lstrip('0')
    return s or '0'


def to_number(v):
    """Цена -> float. Понимает '21 900,00' и текстовые числа; иначе None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace('\xa0', '').replace(' ', '').replace(' ', '').strip()
    s = re.sub(r'[^\d,.\-]', '', s)
    if not s:
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.') if s.rfind(',') > s.rfind('.') else s.replace(',', '')
    else:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def is_green(cell):
    f = cell.fill
    return bool(f and f.patternType and getattr(f.fgColor, 'rgb', None) == GREEN)


def load_price(path):
    """-> (green_map, full_map, sheet_info, collisions, warnings)

    green_map — артикул -> цена только из зелёных ячеек (то, что реально обновляем).
    full_map  — артикул -> цена из всех строк прайса (нужно, чтобы отличить
                «нет в прайсе» от «есть, но не подсвечено»).
    """
    styled = openpyxl.load_workbook(path)              # со стилями — для заливок
    values = openpyxl.load_workbook(path, data_only=True)  # со значениями формул
    green_map, full_map, info, collisions, warns = {}, {}, [], [], []

    for ws in styled.worksheets:
        wsv = values[ws.title]
        hdr = acol = pcol = None
        for r in range(1, min(9, ws.max_row + 1)):
            names = {str(wsv.cell(r, c).value).strip().lower(): c
                     for c in range(1, ws.max_column + 1) if wsv.cell(r, c).value is not None}
            if ART_HEADER in names:
                hdr, acol, pcol = r, names[ART_HEADER], names.get(PRICE_HEADER)
                break
        if hdr is None or pcol is None:
            warns.append(f'Лист «{ws.title}»: не найдены колонки «Артикул»/«ЕРЦ» — лист пропущен.')
            continue

        n_green_rows = n_arts = 0
        for r in range(hdr + 1, ws.max_row + 1):
            raw = wsv.cell(r, acol).value
            if raw is None:
                continue
            arts = [a for a in (norm_article(x) for x in str(raw).split('\n')) if a]
            if not arts:
                continue
            price = to_number(wsv.cell(r, pcol).value)
            if price is None:
                continue
            green = is_green(ws.cell(r, pcol))
            # зелёная заливка вне колонки ЕРЦ — сигнализируем, но не используем
            for c in range(1, ws.max_column + 1):
                if c != pcol and is_green(ws.cell(r, c)):
                    warns.append(f'Лист «{ws.title}», строка {r}: зелёная заливка в колонке '
                                 f'{get_column_letter(c)} (не «ЕРЦ») — проигнорирована.')
            if green:
                n_green_rows += 1
            for a in arts:
                full_map.setdefault(a, (price, ws.title, r))
                if green:
                    if a in green_map and green_map[a][0] != price:
                        collisions.append({'Артикул (норм.)': a,
                                           'Цена 1': green_map[a][0], 'Где 1': f'{green_map[a][1]} r{green_map[a][2]}',
                                           'Цена 2': price, 'Где 2': f'{ws.title} r{r}'})
                        continue
                    green_map[a] = (price, ws.title, r)
                    n_arts += 1
        info.append({'Лист прайса': ws.title, 'Видимость': ws.sheet_state,
                     'Колонка артикула': get_column_letter(acol),
                     'Колонка цены (ЕРЦ)': get_column_letter(pcol),
                     'Зелёных строк': n_green_rows, 'Артикулов из зелёных строк': n_arts})
    return green_map, full_map, info, collisions, warns


def _record(feed, row_no, raw, art, old, green_map, full_map):
    g = green_map.get(art)
    f = full_map.get(art)
    new = g[0] if g else None
    if old is None:
        status, applied = ST_NOPRICE, None
    elif g is None and f is None:
        status, applied = ST_ABSENT, None
    elif g is None:
        status, applied = ST_NOT_GREEN, None
    elif new == old:
        status, applied = ST_GREEN_SAME, None
    else:
        status, applied = ST_UPDATED, new
    delta = pct = None
    if status == ST_UPDATED:
        delta = new - old
        pct = delta / old if old else None
    src = f'{g[1]} r{g[2]}' if g else (f'{f[1]} r{f[2]}' if f else '')
    return {'Партнёр': feed['partner'], 'Файл': feed['file'], 'Лист': feed['sheet'],
            'Строка': row_no, 'Артикул (исходный)': str(raw).strip(),
            'Артикул (норм.)': art, 'Старая цена': old,
            'Новая цена': new, 'Дельта': delta, 'Дельта %': pct,
            'Статус': status, 'Источник в прайсе': src,
            'Цена в прайсе (справочно)': f[0] if f else None}, applied


def process_xls(feed, green_map, full_map, src, dst):
    rb = xlrd.open_workbook(src, formatting_info=True)
    rs = rb.sheet_by_index(feed['sheet_idx'])
    wb = xlcopy(rb)
    ws = wb.get_sheet(feed['sheet_idx'])
    rows = []
    for r in range(feed['header'] + 1, rs.nrows):
        raw = rs.cell_value(r, feed['article'])
        art = norm_article(raw)
        if art is None:
            continue
        old = to_number(rs.cell_value(r, feed['price']))
        rec, applied = _record(feed, r + 1, raw, art, old, green_map, full_map)
        rows.append(rec)
        if applied is not None:
            # xlwt.write() сбрасывает оформление, поэтому запоминаем индекс стиля
            # ячейки в СКОПИРОВАННОЙ книге (у неё своя нумерация XF) и возвращаем его.
            cells = ws.row(r)._Row__cells
            prev = cells.get(feed['price'])
            xf = prev.xf_idx if prev is not None else None
            ws.write(r, feed['price'], applied)
            if xf is not None:
                cells[feed['price']].xf_idx = xf
    wb.save(dst)
    return rows


def process_xlsx(feed, green_map, full_map, src, dst):
    wb = openpyxl.load_workbook(src)
    ws = wb[feed['sheet']]
    rows = []
    for r in range(feed['header'] + 2, ws.max_row + 1):
        raw = ws.cell(r, feed['article'] + 1).value
        art = norm_article(raw)
        if art is None:
            continue
        cell = ws.cell(r, feed['price'] + 1)
        old = to_number(cell.value)
        rec, applied = _record(feed, r, raw, art, old, green_map, full_map)
        rows.append(rec)
        if applied is not None:
            cell.value = applied      # стиль ячейки openpyxl сохраняет сам
    wb.save(dst)
    return rows


# ---------- отчёт ----------
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
UP_FILL = PatternFill('solid', fgColor='FDECEA')
DOWN_FILL = PatternFill('solid', fgColor='E8F4EA')
MONEY = '# ##0'
PCT = '+0.0%;-0.0%;0.0%'


def _dump(wb, title, df, money=(), pct=(), zebra=None, note=None):
    ws = wb.create_sheet(title[:31])
    start = 1
    if note:
        ws.cell(1, 1, note).font = Font(italic=True, color='555555')
        ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[1].height = 30
        start = 3
    if df is None or len(df) == 0:
        c = ws.cell(start, 1, '— пусто —')
        c.font = Font(bold=True, italic=True, color='777777')
        return
    ws.append([]) if False else None
    for j, col in enumerate(df.columns, 1):
        ws.cell(start, j, col)
    for i, row in enumerate(df.itertuples(index=False), start + 1):
        for j, v in enumerate(row, 1):
            ws.cell(i, j, '' if v is None or (isinstance(v, float) and pd.isna(v)) else v)
    for c in ws[start]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = ws.cell(start + 1, 1)
    ws.auto_filter.ref = f'A{start}:{get_column_letter(len(df.columns))}{start + len(df)}'
    for j, col in enumerate(df.columns, 1):
        letter = get_column_letter(j)
        width = max([len(str(col))] + [len(str(v)) for v in df.iloc[:300, j - 1] if v is not None] or [8])
        ws.column_dimensions[letter].width = min(max(width + 2, 11), 44)
        if col in money:
            for i in range(start + 1, start + len(df) + 1):
                ws.cell(i, j).number_format = MONEY
        if col in pct:
            for i in range(start + 1, start + len(df) + 1):
                ws.cell(i, j).number_format = PCT
    if zebra and zebra in df.columns:
        k = list(df.columns).index(zebra) + 1
        for i in range(start + 1, start + len(df) + 1):
            v = ws.cell(i, k).value
            if isinstance(v, (int, float)) and v:
                fill = UP_FILL if v > 0 else DOWN_FILL
                for j in range(1, len(df.columns) + 1):
                    ws.cell(i, j).fill = fill


def build_report(all_rows, green_map, full_map, sheet_info, collisions, warns, dst):
    df = pd.DataFrame(all_rows)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = []
    for partner, g in df.groupby('Партнёр', sort=False):
        summary.append({'Партнёр': partner, 'Файл': g['Файл'].iloc[0], 'Лист': g['Лист'].iloc[0],
                        'Всего позиций': len(g),
                        'Обновлено': int((g['Статус'] == ST_UPDATED).sum()),
                        'Зелёная, цена совпадала': int((g['Статус'] == ST_GREEN_SAME).sum()),
                        'В прайсе, но не зелёная': int((g['Статус'] == ST_NOT_GREEN).sum()),
                        'Нет в прайсе': int((g['Статус'] == ST_ABSENT).sum()),
                        'Без старой цены': int((g['Статус'] == ST_NOPRICE).sum())})
    for s in SKIPPED:
        summary.append({'Партнёр': s['partner'], 'Файл': s['file'], 'Лист': '—', 'Всего позиций': 18,
                        'Обновлено': 0, 'Зелёная, цена совпадала': '—', 'В прайсе, но не зелёная': '—',
                        'Нет в прайсе': '—', 'Без старой цены': 'нет артикулов'})
    _dump(wb, 'Сводка', pd.DataFrame(summary),
          note='Обновлялись только цены, подсвеченные в прайсе зелёным. Прочие поля фидов не изменялись.')

    cols = ['Партнёр', 'Файл', 'Лист', 'Строка', 'Артикул (исходный)', 'Артикул (норм.)',
            'Старая цена', 'Новая цена', 'Дельта', 'Дельта %', 'Источник в прайсе']
    diff = df[df['Статус'] == ST_UPDATED][cols].copy()
    diff = diff.reindex(diff['Дельта %'].abs().sort_values(ascending=False).index)
    _dump(wb, 'Расхождения', diff, money=('Старая цена', 'Новая цена', 'Дельта'),
          pct=('Дельта %',), zebra='Дельта',
          note='Строки, где цена в фиде изменена. Розовый — цена выросла, зелёный — снизилась.')

    absent = df[df['Статус'] == ST_ABSENT][['Партнёр', 'Файл', 'Лист', 'Строка',
                                            'Артикул (исходный)', 'Артикул (норм.)', 'Старая цена']]
    _dump(wb, 'Нет в прайсе', absent, money=('Старая цена',),
          note='Артикулов нет в новом прайсе ни в одной строке. Оставлены со старой ценой.')

    ng = df[df['Статус'] == ST_NOT_GREEN].copy()
    ng['Отличается от прайса'] = ng.apply(
        lambda r: 'да' if r['Цена в прайсе (справочно)'] is not None
        and r['Старая цена'] is not None
        and r['Цена в прайсе (справочно)'] != r['Старая цена'] else 'нет', axis=1)
    ng = ng[['Партнёр', 'Файл', 'Строка', 'Артикул (норм.)', 'Старая цена',
             'Цена в прайсе (справочно)', 'Отличается от прайса', 'Источник в прайсе']]
    ng = ng.sort_values('Отличается от прайса', ascending=False)
    _dump(wb, 'В прайсе, но не зелёные', ng, money=('Старая цена', 'Цена в прайсе (справочно)'),
          note='Не трогали: в прайсе цена не подсвечена зелёным. Колонка «Отличается от прайса» — '
               'справочно, на случай если такие строки тоже нужно обновить.')

    homo = []
    for rec in df.itertuples(index=False):
        a = rec._5  # 'Артикул (норм.)'
        if a in full_map:
            continue
        b = a.translate(CYR_TO_LAT)
        if b != a and b in full_map:
            homo.append({'Партнёр': rec.Партнёр, 'Строка': rec.Строка, 'Артикул в фиде': a,
                         'Совпал бы с прайсом как': b, 'Цена в фиде': rec._6,
                         'Цена в прайсе': full_map[b][0],
                         'Подсвечена зелёным': 'да' if b in green_map else 'нет',
                         'Источник в прайсе': f'{full_map[b][1]} r{full_map[b][2]}'})
    _dump(wb, 'Кириллица в артикуле', pd.DataFrame(homo),
          money=('Цена в фиде', 'Цена в прайсе'),
          note='Артикул не совпал с прайсом только из-за кириллических букв, визуально '
               'неотличимых от латинских (О, С, Р, А…). Цены НЕ менялись. '
               'Если такие позиции нужно сматчить — скажите, добавлю в правило нормализации.')

    used = set(df['Артикул (норм.)'])
    unused = pd.DataFrame([{'Артикул (норм.)': a, 'Новая цена': v[0], 'Источник в прайсе': f'{v[1]} r{v[2]}'}
                           for a, v in sorted(green_map.items()) if a not in used])
    _dump(wb, 'Зелёные — нет в фидах', unused, money=('Новая цена',),
          note='Позиции с новой (зелёной) ценой, которых нет ни в одном обработанном фиде.')

    _dump(wb, 'Разбор прайса', pd.DataFrame(sheet_info),
          note=f'Зелёная заливка #{GREEN[2:]}. Всего артикулов с новой ценой: {len(green_map)}; '
               f'всего артикулов в прайсе: {len(full_map)}.')

    if collisions:
        _dump(wb, 'Конфликты в прайсе', pd.DataFrame(collisions),
              note='Один артикул встречается в зелёных строках с разными ценами. Взято первое значение.')
    if warns:
        _dump(wb, 'Предупреждения', pd.DataFrame({'Предупреждение': warns}))

    _dump(wb, 'Параметры прогона', pd.DataFrame([
        {'Параметр': 'Файл прайса', 'Значение': PRICE_FILE},
        {'Параметр': 'Что обновляли', 'Значение': f'только цены с зелёной заливкой #{GREEN[2:]} в колонке «ЕРЦ»'},
        {'Параметр': 'Что не меняли', 'Значение': 'все остальные поля фидов — без изменений'},
        {'Параметр': 'Нормализация артикула', 'Значение': 'trim, схлопывание пробелов (вкл. NBSP), UPPER, срез ведущих нулей'},
        {'Параметр': 'Многострочные артикулы', 'Значение': 'ячейка «Артикул» разбивается по переводу строки; все артикулы строки получают её цену «ЕРЦ»'},
        {'Параметр': 'Округление', 'Значение': 'не применялось — цена перенесена как в прайсе'},
        {'Параметр': 'Артикулов с новой ценой', 'Значение': len(green_map)},
        {'Параметр': 'Всего позиций обновлено', 'Значение': int((df['Статус'] == ST_UPDATED).sum())},
        {'Параметр': 'М&M', 'Значение': SKIPPED[0]['reason']},
    ]))
    wb.save(dst)
    return df


def main():
    price_path = os.path.join(FEEDS_DIR, PRICE_FILE)
    green_map, full_map, sheet_info, collisions, warns = load_price(price_path)
    print(f'Прайс: артикулов с зелёной ценой {len(green_map)}, всего артикулов {len(full_map)}')
    for w in warns:
        print('  ! ' + w)

    os.makedirs(OUT_DIR, exist_ok=True)
    all_rows = []
    for feed in FEEDS:
        src = os.path.join(FEEDS_DIR, feed['file'])
        dst = os.path.join(OUT_DIR, feed['file'])
        fn = process_xls if feed['kind'] == 'xls' else process_xlsx
        rows = fn(feed, green_map, full_map, src, dst)
        all_rows += rows
        upd = sum(1 for r in rows if r['Статус'] == ST_UPDATED)
        print(f"  {feed['partner']:22s} позиций {len(rows):4d} | обновлено {upd:4d}")
    for s in SKIPPED:
        shutil.copy2(os.path.join(FEEDS_DIR, s['file']), os.path.join(OUT_DIR, s['file']))
        print(f"  {s['partner']:22s} скопирован без изменений")

    rep = os.path.join(OUT_DIR, 'Сверка цен.xlsx')
    build_report(all_rows, green_map, full_map, sheet_info, collisions, warns, rep)
    print(f'Готово: {OUT_DIR}/ + {rep}')


if __name__ == '__main__':
    main()
